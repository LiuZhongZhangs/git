#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
读excel文件，生成edgeconfig.ini
'''
import random
import time
import os
import sys
import openpyxl


from configparser import ConfigParser

#g_edgeconfig_file = "../etc/edge_config.ini"



def gen_configfile(xlsx_file, device_id_input):

    cfg = ConfigParser()

    #读excel文件
    wb=openpyxl.load_workbook(xlsx_file)
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])

    print("sheet name:%s" % sheet.title)
    
    nrows = sheet.max_row
    ncolumns = sheet.max_column
    print("rows:%d, columns:%d" %(nrows, ncolumns))

    if sheet.cell(row=1, column=3).value != "网络类型":
        print("未升级", sheet.cell(row=1, column=2).value)
        return -1
    else
        print("已升级", sheet.cell(row=1, column=2).value)
        return -2

if __name__ == '__main__':
    excel_file = sys.argv[1]
    device_id_input = sys.argv[2]
    ret = gen_configfile(excel_file, device_id_input)

    if ret == -1:
        print("未升级")
        sys.exit(0)
    else:
        print("已升级")
        sys.exit(1)



        
